import os
import sys
import json
import hashlib
import librosa
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, timedelta
from mutagen import File as MutagenFile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
import logging
import shutil
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import queue

# Attempt to include local FFmpeg if present
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOCAL_FFMPEG = os.path.join(SCRIPT_DIR, 'ffmpeg', 'bin')
if os.path.isdir(LOCAL_FFMPEG):
    os.environ['PATH'] = LOCAL_FFMPEG + os.pathsep + os.environ.get('PATH', '')

# Configurations
SUPPORTED_EXTENSIONS = ['.flac', '.aiff', '.aif', '.m4a', '.mp3', '.wav']
MIN_SIZE = 1024  # skip files smaller than 1KB
STATE_FILE = 'processed_state.json'
EXCEL_FILE = 'audio_analysis.xlsx'
LOG_FILE = 'program.log'
THRESH_DB = -60
PROPORTION_THRESHOLD = 0.05
N_FFT = 4096
WEIGHTS = {'freq': 40, 'bitrate': 30, 'samplerate': 20, 'bitdepth': 10}
MAX_WORKERS = 4  # concurrent threads

# Logging setup
logging.basicConfig(filename=LOG_FILE, level=logging.ERROR,
                    format='%(asctime)s %(levelname)s: %(message)s')
logger = logging.getLogger()

# Utility functions

def load_state(path):
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            logger.exception(f"Error loading state from {path}")
    return {}


def save_state(path, state):
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=2, ensure_ascii=False)
    except Exception:
        logger.exception(f"Error saving state to {path}")


def file_hash(path, block_size=65536):
    hasher = hashlib.md5()
    try:
        with open(path, 'rb') as f:
            buf = f.read(block_size)
            while buf:
                hasher.update(buf)
                buf = f.read(block_size)
    except Exception:
        logger.exception(f"Hash error for {path}")
        return None
    return hasher.hexdigest()


def max_reliable_frequency(path):
    try:
        y, sr = librosa.load(path, sr=None, mono=True)
        S = np.abs(librosa.stft(y, n_fft=N_FFT))
        freqs = librosa.fft_frequencies(sr=sr, n_fft=N_FFT)
        mags_db = librosa.amplitude_to_db(S, ref=np.max)
        presence = (mags_db > THRESH_DB).mean(axis=1)
        valid = np.where(presence >= PROPORTION_THRESHOLD)[0]
        max_freq = int(freqs[valid[-1]]) if valid.size else 0
        return max_freq, sr
    except Exception:
        logger.exception(f"Frequency analysis error for {path}")
        return None, None


def extract_metadata(path):
    try:
        f = MutagenFile(path)
        info = f.info
        return getattr(info, 'bitrate', None), getattr(info, 'sample_rate', None), getattr(info, 'bits_per_sample', None)
    except Exception:
        logger.exception(f"Metadata error for {path}")
        return None, None, None


def compute_rating(freq, sr, bitrate, bitdepth):
    freq_score = min(freq / 20000.0, 1.0) * WEIGHTS['freq']
    br_score = min(bitrate / 320000.0, 1.0) * WEIGHTS['bitrate'] if bitrate else 0
    sr_score = min(sr / 48000.0, 1.0) * WEIGHTS['samplerate'] if sr else 0
    bd_score = min(bitdepth / 24.0, 1.0) * WEIGHTS['bitdepth'] if bitdepth else 0
    total_w = sum(WEIGHTS.values())
    raw = freq_score + br_score + sr_score + bd_score
    return int(raw / total_w * 100) if total_w else 0


def needs_reanalysis(entry):
    if not entry:
        return True
    for key in ('freq', 'duration', 'bitrate', 'samplerate'):
        if str(entry.get(key)) in ('ERROR', 'N/A', 'None'):
            return True
    return False


def write_excel(state):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Audio Report'

        headers = [
            'file_name', 'file_size_bytes', 'duration_s',
            'max_freq_hz', 'bitrate', 'samplerate',
            'bitdepth', 'rating'
        ]
        ws.append(headers)

        for e in state.values():
            ws.append([
                os.path.basename(e['path']),
                e['size'],
                e['duration'],
                e['freq'],
                e['bitrate'],
                e['samplerate'],
                e['bitdepth'],
                e['rating']
            ])

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for col in ws.columns:
            max_len = max(len(str(c.value)) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        rating_range = f"H2:H{ws.max_row}"

        low_rule = CellIsRule(
            operator='lessThan',
            formula=['50'],
            stopIfTrue=True,
            fill=PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        )
        ws.conditional_formatting.add(rating_range, low_rule)

        mid_rule = CellIsRule(
            operator='between',
            formula=['50', '79'],
            stopIfTrue=True,
            fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        )
        ws.conditional_formatting.add(rating_range, mid_rule)

        good_rule = CellIsRule(
            operator='between',
            formula=['80', '89'],
            stopIfTrue=True,
            fill=PatternFill(start_color='66CC66', end_color='66CC66', fill_type='solid')
        )
        ws.conditional_formatting.add(rating_range, good_rule)

        high_rule = CellIsRule(
            operator='greaterThanOrEqual',
            formula=['90'],
            stopIfTrue=True,
            fill=PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        )
        ws.conditional_formatting.add(rating_range, high_rule)

        error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        ws.conditional_formatting.add(
            f"C2:C{ws.max_row}",
            CellIsRule(operator='equal', formula=['\"ERROR\"'], fill=error_fill)
        )
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            CellIsRule(operator='equal', formula=['\"ERROR\"'], fill=error_fill)
        )

        wb.save(EXCEL_FILE)
        return True

    except Exception:
        logger.exception(f"Excel save error: {EXCEL_FILE}")
        return False


class AnalyzerApp:
    def __init__(self, master):
        self.master = master
        master.title('Audio Frequency Analyzer')
        master.geometry('900x600')

        # Check for FFmpeg
        if not shutil.which('ffmpeg'):
            messagebox.showwarning(
                'FFmpeg Not Found',
                'FFmpeg was not found in PATH. For full format support, please install FFmpeg.'
            )

        # Format selection
        self.selected_exts = set(SUPPORTED_EXTENSIONS)
        fmt_btn = ttk.Button(master, text='Select Formats', command=self.open_format_window)
        fmt_btn.pack(anchor='ne', padx=10, pady=5)

        # Layout frames
        top = ttk.Frame(master); top.pack(fill=tk.X, padx=5, pady=5)
        mid = ttk.Frame(master); mid.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        bot = ttk.Frame(master); bot.pack(fill=tk.X, padx=5, pady=5)

        # Controls
        select_btn = ttk.Button(top, text='Select Folder', command=self.select_folder)
        select_btn.pack(side=tk.LEFT)
        self.progress = ttk.Progressbar(top, length=400, mode='determinate')
        self.progress.pack(side=tk.LEFT, padx=10)
        self.remaining_label = ttk.Label(top, text='Remaining: calculating...')
        self.remaining_label.pack(side=tk.LEFT); self.remaining_label.pack_forget()

        # Current tasks
        cf = ttk.LabelFrame(mid, text='Currently Processing'); cf.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        self.current_list = tk.Listbox(cf); self.current_list.pack(fill=tk.BOTH, expand=True)

        # Results
        rf = ttk.LabelFrame(mid, text='Results'); rf.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        cols = ('file','freq','bitrate','samplerate','bitdepth','rating')
        self.tree = ttk.Treeview(rf, columns=cols, show='headings')
        for c in cols: self.tree.heading(c,text=c); self.tree.column(c,width=100)
        self.tree.pack(fill=tk.BOTH, expand=True)

        # Internal state
        self.queue = queue.Queue()
        self.start_time = None
        self.processed = 0
        self.total = 0
        self.samples_needed = 5

    def open_format_window(self):
        win = tk.Toplevel(self.master); win.title('Select Formats')
        vars_ = {}
        tk.Label(win, text='Choose extensions:').pack(anchor='w', padx=10, pady=5)
        for ext in SUPPORTED_EXTENSIONS:
            var = tk.BooleanVar(value=True); cb = tk.Checkbutton(win, text=ext, variable=var)
            cb.pack(anchor='w', padx=20); vars_[ext] = var
        def on_ok(): self.selected_exts = {e for e,v in vars_.items() if v.get()}; win.destroy()
        ttk.Button(win, text='OK', command=on_ok).pack(pady=10)

    def select_folder(self):
        folder = filedialog.askdirectory();
        if not folder: return
        files = [os.path.join(r,f) for r,_,fs in os.walk(folder) for f in fs
                 if os.path.splitext(f)[1].lower() in self.selected_exts and os.path.getsize(os.path.join(r,f))>=MIN_SIZE]
        if not files:
            messagebox.showwarning('No files','No supported files found.'); return
        self.total=len(files); self.progress['maximum']=self.total; self.start_time=time.time(); self.processed=0
        self.remaining_label.pack_forget()
        self.executor=ThreadPoolExecutor(max_workers=MAX_WORKERS)
        self.futures={self.executor.submit(self.process_file,f):f for f in files}
        threading.Thread(target=self.collect_results,args=(self.futures,),daemon=True).start()

    def collect_results(self,futures):
        for future in as_completed(futures): self.queue.put(('done',future.result()))
        self.queue.put(('all_done',None))

    def process_file(self,filepath):
        self.queue.put(('start',filepath))
        stat=os.stat(filepath); size,mtime=stat.st_size,stat.st_mtime
        h=file_hash(filepath); entry=state.get(h)
        if entry and entry['mtime']==mtime and entry['size']==size and not needs_reanalysis(entry): return entry
        freq,sr=max_reliable_frequency(filepath); br,sr_meta,bd=extract_metadata(filepath)
        try: duration=int(librosa.get_duration(path=filepath))
        except: duration='ERROR'
        rating=compute_rating(freq or 0,sr or 0,br,bd)
        entry={'path':filepath,'size':size,'mtime':mtime,'duration':duration,'freq':freq or 0,'bitrate':br or 'N/A','samplerate':sr_meta or 'N/A','bitdepth':bd or 'N/A','rating':rating}
        state[h]=entry; return entry

    def update_ui(self):
        try:
            while True:
                action,data=self.queue.get_nowait()
                if action=='start': self.current_list.insert(tk.END,os.path.basename(data))
                elif action=='done':
                    self.processed+=1
                    try: idx=self.current_list.get(0,tk.END).index(os.path.basename(data['path'])); self.current_list.delete(idx)
                    except: pass
                    self.tree.insert('',tk.END,values=(os.path.basename(data['path']),data['freq'],data['bitrate'],data['samplerate'],data['bitdepth'],data['rating']))
                    self.progress['value']=self.processed
                    if self.processed>=self.samples_needed:
                        elapsed=time.time()-self.start_time; rate=elapsed/self.processed; rem=self.total-self.processed; rem_sec=int(rem*rate)
                        mins,secs=divmod(rem_sec,60)
                        self.remaining_label.config(text=f'Remaining: {mins}m{secs}s'); self.remaining_label.pack(side=tk.LEFT)
                elif action=='all_done':
                    save_state(STATE_FILE,state)
                    if write_excel(state): messagebox.showinfo('Done','Analysis and report generated successfully.')
                    else: messagebox.showerror('Error','Failed to save Excel report. JSON state saved; you can regenerate Excel later.')
                    self.master.destroy()
                self.queue.task_done()
        except queue.Empty:
            pass
        finally:
            if self.master.winfo_exists(): self.master.after(100,self.update_ui)

if __name__=='__main__':
    state=load_state(STATE_FILE)
    root=tk.Tk(); app=AnalyzerApp(root); root.after(100,app.update_ui); root.mainloop()
