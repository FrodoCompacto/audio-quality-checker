import os
import sys
import json
import hashlib
import librosa
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
from mutagen import File as MutagenFile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
import logging
import traceback

# Attempt to include local FFmpeg (portable standalone) if present
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOCAL_FFMPEG = os.path.join(SCRIPT_DIR, 'ffmpeg', 'bin')
if os.path.isdir(LOCAL_FFMPEG):
    os.environ['PATH'] = LOCAL_FFMPEG + os.pathsep + os.environ.get('PATH', '')

# Verifica backend do FFmpeg
import shutil
if not shutil.which('ffmpeg'):
    # Informar apenas no console, sem interromper; audioread pode falhar depois
    print("[WARNING] FFmpeg não encontrado no PATH. Para melhor compatibilidade com MP3/M4A, coloque um build standalone em './ffmpeg/bin' ou instale FFmpeg globalmente.")

# Configurações
SUPPORTED_EXTENSIONS = ('.flac', '.aiff', '.aif', '.m4a', '.mp3', '.wav')
THRESH_DB = -60
PROPORTION_THRESHOLD = 0.05
N_FFT = 4096
STATE_FILE = 'processed_state.json'
EXCEL_FILE = 'audio_analysis.xlsx'
LOG_FILE = 'program.log'

WEIGHTS = {
    'freq': 40,
    'bitrate': 30,
    'samplerate': 20,
    'bitdepth': 10
}

# Configurar logging
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.DEBUG,
    format='%(asctime)s %(levelname)s: %(message)s'
)
logger = logging.getLogger()

# ... rest of code unchanged ...

def load_state(path):
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            logger.exception(f"Erro ao carregar estado de {path}")
            return {}
    return {}


def save_state(path, state):
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=2, ensure_ascii=False)
        logger.info(f"Estado salvo em {path}")
    except Exception:
        logger.exception(f"Erro ao salvar estado em {path}")


def file_hash(path, block_size=65536):
    hasher = hashlib.md5()
    try:
        with open(path, 'rb') as f:
            buf = f.read(block_size)
            while buf:
                hasher.update(buf)
                buf = f.read(block_size)
    except Exception:
        logger.exception(f"Hash generation error for {path}")
        return None
    return hasher.hexdigest()


def max_reliable_frequency(path):
    try:
        y, sr = librosa.load(path, sr=None, mono=True)
        S = np.abs(librosa.stft(y, n_fft=N_FFT))
        freqs = librosa.fft_frequencies(sr=sr, n_fft=N_FFT)
        mags_db = librosa.amplitude_to_db(S, ref=np.max)
        presence = (mags_db > THRESH_DB).mean(axis=1)
        valid_bins = np.where(presence >= PROPORTION_THRESHOLD)[0]
        max_freq = float(freqs[valid_bins[-1]]) if valid_bins.size else 0.0
        return max_freq, sr
    except Exception:
        logger.exception(f"Erro ao analisar frequência para {path}")
        return None, None


def extract_metadata(filepath):
    try:
        f = MutagenFile(filepath)
        info = f.info
        bitrate = getattr(info, 'bitrate', None)
        samplerate = getattr(info, 'sample_rate', None)
        bitdepth = getattr(info, 'bits_per_sample', None)
        return bitrate, samplerate, bitdepth
    except Exception:
        logger.exception(f"Metadata extraction error for {filepath}")
        return None, None, None


def compute_rating(freq, sr, bitrate, bitdepth):
    freq_score = min(freq / 20000.0, 1.0) * WEIGHTS['freq']
    br_score = min(bitrate / 320000.0, 1.0) * WEIGHTS['bitrate'] if bitrate else 0
    sr_score = min(sr / 48000.0, 1.0) * WEIGHTS['samplerate'] if sr else 0
    bd_score = min(bitdepth / 24.0, 1.0) * WEIGHTS['bitdepth'] if bitdepth else 0
    return round(freq_score + br_score + sr_score + bd_score, 1)


def write_excel(state):
    wb = Workbook()
    ws = wb.active
    ws.title = "Audio Report"

    headers = ['file_name', 'file_size_bytes', 'duration_s', 'max_freq_hz',
               'bitrate', 'samplerate', 'bitdepth', 'rating']
    ws.append(headers)

    for e in state.values():
        filename = os.path.basename(e['path'])
        ws.append([
            filename,
            e['size'],
            round(e['duration'], 2) if isinstance(e['duration'], float) else e['duration'],
            e['freq'],
            e['bitrate'],
            e['samplerate'],
            e['bitdepth'],
            e['rating']
        ])

    # Formatação
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        length = max(len(str(c.value)) for c in col)
        ws.column_dimensions[col[0].column_letter].width = length + 2

    # Formatação condicional
    rating_col = 'H'
    ws.conditional_formatting.add(
        f"{rating_col}2:{rating_col}{ws.max_row}",
        ColorScaleRule(start_type='num', start_value=0, start_color='FF0000',
                       mid_type='num', mid_value=50, mid_color='FFFF00',
                       end_type='num', end_value=100, end_color='00FF00')
    )

    error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    for col in ['C', 'D']:
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            CellIsRule(operator='equal', formula=['"ERROR"'], fill=error_fill)
        )

    # Tentar salvar o Excel
    try:
        wb.save(EXCEL_FILE)
        logger.info(f"Excel salvo: {EXCEL_FILE}")
        return True
    except Exception:
        logger.exception(f"Falha ao salvar Excel: {EXCEL_FILE}")
        return False


def scan_and_update(root_dir):
    state = load_state(STATE_FILE)
    updated = False

    files = [os.path.join(dp, f) for dp, _, files in os.walk(root_dir) for f in files
             if f.lower().endswith(SUPPORTED_EXTENSIONS)]
    total = len(files)
    count = 0

    for filepath in files:
        count += 1
        try:
            stat = os.stat(filepath)
        except Exception:
            logger.exception(f"Access error for {filepath}")
            print(f"[{count}/{total}] Access error: {filepath}")
            continue

        h = file_hash(filepath)
        if not h:
            print(f"[{count}/{total}] Hash error: {filepath}")
            continue

        mtime, size = stat.st_mtime, stat.st_size
        entry = state.get(h)
        if entry and entry.get('mtime') == mtime and entry.get('size') == size:
            print(f"[{count}/{total}] Already analyzed: {filepath}")
            continue

        freq, sr = max_reliable_frequency(filepath)
        bitrate, samplerate_meta, bitdepth = extract_metadata(filepath)
        try:
            duration = librosa.get_duration(filename=filepath)
        except Exception:
            logger.exception(f"Duration error for {filepath}")
            duration = 'ERROR'

        rating = compute_rating(freq or 0, sr or 0, bitrate, bitdepth)
        state[h] = {
            'path': filepath,
            'mtime': mtime,
            'size': size,
            'duration': duration,
            'freq': freq if freq is not None else 'ERROR',
            'bitrate': bitrate or 'N/A',
            'samplerate': samplerate_meta or 'N/A',
            'bitdepth': bitdepth or 'N/A',
            'rating': rating
        }
        updated = True
        print(f"[{count}/{total}] Analyzed: {filepath} → {freq or 'ERROR'} Hz, Rating: {rating}%")

    if updated:
        # Tentar escrever o Excel antes de salvar estado
        if write_excel(state):
            save_state(STATE_FILE, state)
            messagebox.showinfo('Done', 'Analysis and report generated successfully.')
        else:
            messagebox.showerror(
                'Error',
                'Failed to save Excel report. This probably happened because the Excel report is open.\n'
                'Please close the file or any software using it and try again.\n'
                'If the problem persists, check the log file at program.log.'
            )
    else:
        print("No new or updated tracks found.")
        messagebox.showinfo('Info', 'No new or updated tracks found.')


def select_folder_and_run():
    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title='Select audio folder')
    if not folder:
        messagebox.showinfo('Cancelled', 'No folder selected.')
        return
    scan_and_update(folder)


if __name__ == '__main__':
    select_folder_and_run()
